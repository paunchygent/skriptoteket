"""Openpyxl renderer for classroom-planner seating XLSX exports.

Purpose:
    Generate the teacher-facing seating workbook locally inside Skriptoteket so
    seating XLSX exports keep the classroom's spatial layout without depending
    on Sir Convert-a-Lot.

Relationships:
    - Implements `SeatingXlsxRendererProtocol`.
    - Consumes the application-layer seating XLSX workbook view model.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from skriptoteket.application.curated_apps.classroom_planner.exports import (
    seating_xlsx_view_model,
)
from skriptoteket.protocols.classroom_planner_exports import SeatingXlsxRendererProtocol

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SECTION_FILL = PatternFill(fill_type="solid", fgColor="E5E7EB")
_SECTION_FONT = Font(size=12, bold=True)
_ASSIGNED_SEAT_FILL = PatternFill(fill_type="solid", fgColor="DBEAFE")
_EMPTY_SEAT_FILL = PatternFill(fill_type="solid", fgColor="F3F4F6")
_SEAT_SIDE = Side(style="thin", color="475569")
_SEAT_BORDER = Border(
    left=_SEAT_SIDE,
    right=_SEAT_SIDE,
    top=_SEAT_SIDE,
    bottom=_SEAT_SIDE,
)
_SECTION_BORDER = Border(bottom=Side(style="thin", color="1F2937"))
_GRID_START_ROW = 1
_GRID_START_COLUMN = 1
_SEAT_COLUMN_WIDTH = 13
_SPACER_COLUMN_WIDTH = 13
_SEAT_ROW_HEIGHT = 72
_SPACER_ROW_HEIGHT = 72


class SeatingXlsxRenderer(SeatingXlsxRendererProtocol):
    """Render the seating XLSX workbook bytes for one seating draft."""

    def render(self, *, view_model: seating_xlsx_view_model.SeatingXlsxWorkbookViewModel) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise RuntimeError("Workbook did not create an active worksheet.")
        sheet.title = "Sittplacering"
        self._render_sheet(sheet, view_model=view_model)
        workbook.active = 0

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _render_sheet(
        self,
        worksheet,
        *,
        view_model: seating_xlsx_view_model.SeatingXlsxWorkbookViewModel,
    ) -> None:
        self._render_grid(
            worksheet,
            grid_rows=view_model.grid_rows,
            show_seat_comments=True,
        )
        if view_model.unplaced_student_names:
            self._render_unplaced_section(
                worksheet,
                start_row=_unplaced_section_start_row(grid_rows=view_model.grid_rows),
                student_names=view_model.unplaced_student_names,
            )
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

    def _render_grid(
        self,
        worksheet,
        *,
        grid_rows: tuple[seating_xlsx_view_model.SeatingXlsxGridRow, ...],
        show_seat_comments: bool,
    ) -> None:
        if not grid_rows:
            return

        column_count = len(grid_rows[0].cells)
        for column_offset in range(column_count):
            sheet_column = _GRID_START_COLUMN + column_offset
            if _column_contains_seat(grid_rows=grid_rows, column_offset=column_offset):
                worksheet.column_dimensions[_column_letter(sheet_column)].width = _SEAT_COLUMN_WIDTH
            else:
                worksheet.column_dimensions[
                    _column_letter(sheet_column)
                ].width = _SPACER_COLUMN_WIDTH

        for row_offset, grid_row in enumerate(grid_rows):
            sheet_row = _GRID_START_ROW + row_offset
            if any(cell is not None for cell in grid_row.cells):
                worksheet.row_dimensions[sheet_row].height = _SEAT_ROW_HEIGHT
            else:
                worksheet.row_dimensions[sheet_row].height = _SPACER_ROW_HEIGHT
            for column_offset, cell in enumerate(grid_row.cells):
                if cell is None:
                    continue
                worksheet_cell = worksheet.cell(
                    row=sheet_row,
                    column=_GRID_START_COLUMN + column_offset,
                )
                worksheet_cell.value = cell.student_name or ""
                worksheet_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                worksheet_cell.font = Font(
                    bold=cell.student_name is not None,
                    color="0F172A" if cell.student_name is not None else "64748B",
                )
                worksheet_cell.fill = (
                    _ASSIGNED_SEAT_FILL if cell.student_name is not None else _EMPTY_SEAT_FILL
                )
                worksheet_cell.border = _SEAT_BORDER
                if show_seat_comments:
                    worksheet_cell.comment = Comment(cell.seat_label, "Skriptoteket")

    def _render_unplaced_section(
        self,
        worksheet,
        *,
        start_row: int,
        student_names: tuple[str, ...],
    ) -> None:
        worksheet[f"A{start_row}"] = "Ej placerade elever"
        worksheet[f"A{start_row}"].font = _SECTION_FONT
        worksheet[f"A{start_row}"].fill = _SECTION_FILL
        worksheet[f"A{start_row}"].border = _SECTION_BORDER
        worksheet[f"A{start_row + 1}"] = "Elevnamn"
        self._style_header_cell(worksheet[f"A{start_row + 1}"])
        for offset, student_name in enumerate(student_names, start=2):
            worksheet[f"A{start_row + offset}"] = student_name
        worksheet.column_dimensions["A"].width = max(
            worksheet.column_dimensions["A"].width or 0,
            24,
        )

    def _style_header_cell(self, worksheet_cell) -> None:
        worksheet_cell.fill = _HEADER_FILL
        worksheet_cell.font = _HEADER_FONT
        worksheet_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet_cell.border = _SECTION_BORDER


def _column_contains_seat(
    *,
    grid_rows: tuple[seating_xlsx_view_model.SeatingXlsxGridRow, ...],
    column_offset: int,
) -> bool:
    """Return whether any workbook grid row uses the given worksheet column."""

    return any(grid_row.cells[column_offset] is not None for grid_row in grid_rows)


def _column_letter(column_index: int) -> str:
    """Convert a one-based worksheet column index into an Excel column label."""

    label = ""
    current = column_index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        label = chr(65 + remainder) + label
    return label


def _unplaced_section_start_row(
    *,
    grid_rows: tuple[seating_xlsx_view_model.SeatingXlsxGridRow, ...],
) -> int:
    """Place the unplaced section directly below the grid, or at A1 when no seats exist."""

    if not grid_rows:
        return _GRID_START_ROW
    return _GRID_START_ROW + len(grid_rows) + 2
