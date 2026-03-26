"""Openpyxl renderer for classroom-planner grouping XLSX exports.

Purpose:
    Generate the teacher-facing grouping workbook locally inside Skriptoteket
    so the default grouping export artifact is editable while still remaining
    presentation-ready for sharing or PDF printing.

Relationships:
    - Implements `GroupingXlsxRendererProtocol`.
    - Consumes the application-layer grouping XLSX workbook view model.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from skriptoteket.application.curated_apps.classroom_planner.exports import (
    grouping_xlsx_view_model,
)
from skriptoteket.protocols.classroom_planner_exports import GroupingXlsxRendererProtocol

_EDIT_SHEET_TITLE = "Redigera grupper"
_PRESENTATION_SHEET_TITLE = "Dela och exportera"
_HELPER_SHEET_TITLE = "_PresentationData"
_STUDENT_HEADERS = ("Nr i grupp", "Elev", "Grupp (välj)")
_REGISTRY_HEADERS = ("Grupp", "Gruppordning (välj)")
_PRESENTATION_START_ROW = 5
_PRESENTATION_ROW_TYPE_COLUMN = "Z"
_PRESENTATION_GROUP_COLUMN = "AA"
_PRESENTATION_NUMBER_COLUMN = "AB"
_PRESENTATION_STUDENT_COLUMN = "AC"
_TITLE_FONT = Font(size=16, bold=True)
_META_FONT = Font(size=11, italic=True, color="475569")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SECTION_FILL = PatternFill(fill_type="solid", fgColor="E5E7EB")
_SECTION_FONT = Font(size=12, bold=True)
_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
_THIN_SIDE = Side(style="thin", color="CBD5E1")
_PRESENTATION_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)
_EDIT_HELP_TITLE_FONT = Font(size=12, bold=True)
_EDIT_HELP_BODY_FONT = Font(size=10, color="334155")
_EDITABLE_FILL = PatternFill(fill_type="solid", fgColor="FEF3C7")


class GroupingXlsxRenderer(GroupingXlsxRendererProtocol):
    """Render the grouping XLSX workbook bytes for one grouping draft."""

    def render(
        self,
        *,
        view_model: grouping_xlsx_view_model.GroupingXlsxWorkbookViewModel,
    ) -> bytes:
        workbook = Workbook()
        workbook.calculation = CalcProperties(
            calcMode="auto",
            fullCalcOnLoad=True,
            forceFullCalc=True,
        )
        edit_sheet = workbook.active
        if edit_sheet is None:
            raise RuntimeError("Workbook did not create an active worksheet.")
        edit_sheet.title = _EDIT_SHEET_TITLE
        presentation_sheet = workbook.create_sheet(_PRESENTATION_SHEET_TITLE)
        helper_sheet = workbook.create_sheet(_HELPER_SHEET_TITLE)

        self._render_edit_sheet(edit_sheet, view_model=view_model)
        self._render_helper_sheet(helper_sheet, view_model=view_model)
        self._render_presentation_sheet(presentation_sheet, view_model=view_model)

        helper_sheet.sheet_state = "hidden"
        workbook.active = workbook.sheetnames.index(_PRESENTATION_SHEET_TITLE)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _render_edit_sheet(self, worksheet, *, view_model) -> None:
        for column_index, header in enumerate(_STUDENT_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for column_index, header in enumerate(_REGISTRY_HEADERS, start=6):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")

        for row_index, edit_row in enumerate(view_model.edit_rows, start=2):
            worksheet.cell(row=row_index, column=1, value=edit_row.member_order)
            worksheet.cell(row=row_index, column=2, value=edit_row.student_name)
            worksheet.cell(row=row_index, column=3, value=edit_row.group_label)
        for row_index, registry_row in enumerate(view_model.registry_rows, start=2):
            worksheet.cell(row=row_index, column=6, value=registry_row.group_label)
            worksheet.cell(row=row_index, column=7, value=registry_row.group_order)

        worksheet.freeze_panes = "A2"
        worksheet.column_dimensions["A"].width = 12
        worksheet.column_dimensions["B"].width = 28
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["F"].width = 18
        worksheet.column_dimensions["G"].width = 16
        worksheet.column_dimensions["I"].width = 22
        worksheet.column_dimensions["J"].width = 52

        worksheet["I1"] = "Ändra bara detta"
        worksheet["I1"].font = _EDIT_HELP_TITLE_FONT
        worksheet["I2"] = "Kolumn C"
        worksheet["J2"] = "Flytta elever mellan grupper med listan. Skriv inte egna gruppnamn."
        worksheet["I4"] = "Gruppregister"
        worksheet["J4"] = (
            "Ändra bara gruppordningen i kolumn G. Exportera en ny fil för större ändringar."
        )
        worksheet["I5"] = "Inte här"
        worksheet["J5"] = "Lägg inte till eller ta bort elever i Excel-filen."
        for cell_reference in ("I2", "I4", "I5"):
            worksheet[cell_reference].font = Font(bold=True)
        for cell_reference in ("J2", "J4", "J5"):
            worksheet[cell_reference].font = _EDIT_HELP_BODY_FONT
            worksheet[cell_reference].alignment = Alignment(wrap_text=True, vertical="top")

        if view_model.edit_rows:
            table = Table(
                displayName="tblRedigeraGrupper",
                ref=f"A1:C{len(view_model.edit_rows) + 1}",
            )
            table.tableStyleInfo = _TABLE_STYLE
            worksheet.add_table(table)
            if view_model.registry_rows:
                group_validation = DataValidation(
                    type="list",
                    formula1=f"=$F$2:$F${len(view_model.registry_rows) + 1}",
                    allow_blank=True,
                )
                worksheet.add_data_validation(group_validation)
                group_validation.add(f"C2:C{len(view_model.edit_rows) + 1}")
        if view_model.registry_rows:
            table = Table(
                displayName="tblGruppregister",
                ref=f"F1:G{len(view_model.registry_rows) + 1}",
            )
            table.tableStyleInfo = _TABLE_STYLE
            worksheet.add_table(table)
            order_values = ",".join(
                str(index) for index in range(1, len(view_model.registry_rows) + 1)
            )
            order_validation = DataValidation(
                type="list",
                formula1=f'"{order_values}"',
                allow_blank=False,
            )
            worksheet.add_data_validation(order_validation)
            order_validation.add(f"G2:G{len(view_model.registry_rows) + 1}")

        for row_index in range(2, len(view_model.edit_rows) + 2):
            worksheet[f"C{row_index}"].fill = _EDITABLE_FILL
            worksheet[f"C{row_index}"].protection = Protection(locked=False)
        for row_index in range(2, len(view_model.registry_rows) + 2):
            worksheet[f"G{row_index}"].fill = _EDITABLE_FILL
            worksheet[f"G{row_index}"].protection = Protection(locked=False)

        worksheet.protection.sheet = True
        worksheet.protection.sort = False
        worksheet.protection.autoFilter = False
        worksheet.protection.insertRows = False
        worksheet.protection.insertColumns = False
        worksheet.protection.deleteRows = False
        worksheet.protection.deleteColumns = False
        worksheet.protection.formatCells = False
        worksheet.protection.formatColumns = False
        worksheet.protection.formatRows = False

    def _render_helper_sheet(self, worksheet, *, view_model) -> None:
        edit_row_count = len(view_model.edit_rows)
        registry_row_count = len(view_model.registry_rows)
        source_end = edit_row_count + 1
        registry_end = registry_row_count + 1
        sorted_students_end = view_model.sorted_row_capacity + 1
        sorted_registry_end = registry_row_count + 1
        presentation_end = view_model.presentation_row_capacity + 1

        self._write_helper_headers(worksheet)

        for row_index in range(2, source_end + 1):
            worksheet[f"A{row_index}"] = f"='{_EDIT_SHEET_TITLE}'!C{row_index}"
            worksheet[f"B{row_index}"] = (
                f'=IF(A{row_index}="","",IFERROR(INDEX(\'{_EDIT_SHEET_TITLE}\'!$G$2:$G${registry_end},'
                f"MATCH(A{row_index},'{_EDIT_SHEET_TITLE}'!$F$2:$F${registry_end},0)),\"\"))"
            )
            worksheet[f"C{row_index}"] = (
                f'=IF(A{row_index}="",999999,IF(\'{_EDIT_SHEET_TITLE}\'!A{row_index}="",999999,'
                f"'{_EDIT_SHEET_TITLE}'!A{row_index}))"
            )
            worksheet[f"D{row_index}"] = f"='{_EDIT_SHEET_TITLE}'!B{row_index}"
            worksheet[f"E{row_index}"] = f'=IF(OR(A{row_index}="",B{row_index}=""),0,1)'
            worksheet[f"F{row_index}"] = (
                f'=IF(E{row_index}=0,"",1+'
                f'COUNTIFS($E$2:$E${source_end},1,$B$2:$B${source_end},"<"&B{row_index})+'
                f"COUNTIFS($E$2:$E${source_end},1,$B$2:$B${source_end},B{row_index},"
                f'$A$2:$A${source_end},"<"&A{row_index})+'
                f"COUNTIFS($E$2:$E${source_end},1,$B$2:$B${source_end},B{row_index},"
                f'$A$2:$A${source_end},A{row_index},$C$2:$C${source_end},"<"&C{row_index})+'
                f"COUNTIFS($E$2:$E${source_end},1,$B$2:$B${source_end},B{row_index},"
                f"$A$2:$A${source_end},A{row_index},$C$2:$C${source_end},C{row_index},"
                f'$D$2:$D${source_end},"<"&D{row_index}))'
            )

        for row_index in range(2, registry_end + 1):
            worksheet[f"N{row_index}"] = f"='{_EDIT_SHEET_TITLE}'!F{row_index}"
            worksheet[f"O{row_index}"] = f"='{_EDIT_SHEET_TITLE}'!G{row_index}"
            worksheet[f"P{row_index}"] = (
                f'=IF(N{row_index}="","",1+'
                f'COUNTIFS($O$2:$O${registry_end},"<"&O{row_index})+'
                f'COUNTIFS($O$2:$O${registry_end},O{row_index},$N$2:$N${registry_end},"<"&N{row_index}))'
            )

        for row_index in range(2, sorted_registry_end + 1):
            slot_index = row_index - 1
            worksheet[f"R{row_index}"] = (
                f'=IFERROR(INDEX($N$2:$N${registry_end},MATCH({slot_index},$P$2:$P${registry_end},0)),"")'
            )
            worksheet[f"S{row_index}"] = (
                f'=IFERROR(INDEX($O$2:$O${registry_end},MATCH({slot_index},$P$2:$P${registry_end},0)),"")'
            )
            worksheet[f"T{row_index}"] = (
                f'=IF(R{row_index}="","",COUNTIF($H$2:$H${sorted_students_end},R{row_index}))'
            )
            worksheet[f"U{row_index}"] = (
                f'=IF(R{row_index}="","",IF(ROW()=2,1,U{row_index - 1}+T{row_index - 1}+3))'
            )
            worksheet[f"V{row_index}"] = f'=IF(U{row_index}="","",U{row_index}+1)'
            worksheet[f"W{row_index}"] = f'=IF(R{row_index}="","",""&R{row_index}&"|"&S{row_index})'

        for row_index in range(2, sorted_students_end + 1):
            slot_index = row_index - 1
            worksheet[f"H{row_index}"] = (
                f'=IFERROR(INDEX($A$2:$A${source_end},MATCH({slot_index},$F$2:$F${source_end},0)),"")'
            )
            worksheet[f"I{row_index}"] = (
                f'=IFERROR(INDEX($B$2:$B${source_end},MATCH({slot_index},$F$2:$F${source_end},0)),"")'
            )
            worksheet[f"J{row_index}"] = (
                f'=IFERROR(INDEX($C$2:$C${source_end},MATCH({slot_index},$F$2:$F${source_end},0)),"")'
            )
            worksheet[f"K{row_index}"] = (
                f'=IFERROR(INDEX($D$2:$D${source_end},MATCH({slot_index},$F$2:$F${source_end},0)),"")'
            )
            worksheet[f"L{row_index}"] = (
                f'=IF(H{row_index}="","",COUNTIF($H$2:H{row_index},H{row_index}))'
            )
            worksheet[f"M{row_index}"] = (
                f'=IF(H{row_index}="","",IFERROR('
                f'INDEX($U$2:$U${sorted_registry_end},MATCH(H{row_index}&"|"&I{row_index},$W$2:$W${sorted_registry_end},0))'
                f'+L{row_index}+1,""))'
            )

        for row_index in range(2, presentation_end + 1):
            slot_index = row_index - 1
            worksheet[f"Y{row_index}"] = slot_index
            worksheet[f"Z{row_index}"] = (
                f'=IF(COUNTIF($U$2:$U${sorted_registry_end},Y{row_index})>0,"heading",'
                f'IF(COUNTIF($V$2:$V${sorted_registry_end},Y{row_index})>0,"header",'
                f'IF(COUNTIF($M$2:$M${sorted_students_end},Y{row_index})>0,"member","")))'
            )
            worksheet[f"AA{row_index}"] = (
                f'=IF(Z{row_index}="heading",'
                f'IFERROR(INDEX($R$2:$R${sorted_registry_end},MATCH(Y{row_index},$U$2:$U${sorted_registry_end},0)),""),'
                f'IF(Z{row_index}="member",'
                f'IFERROR(INDEX($H$2:$H${sorted_students_end},MATCH(Y{row_index},$M$2:$M${sorted_students_end},0)),""),""))'
            )
            worksheet[f"AB{row_index}"] = (
                f'=IF(Z{row_index}="member",'
                f'IFERROR(INDEX($J$2:$J${sorted_students_end},MATCH(Y{row_index},$M$2:$M${sorted_students_end},0)),""),"")'
            )
            worksheet[f"AC{row_index}"] = (
                f'=IF(Z{row_index}="member",'
                f'IFERROR(INDEX($K$2:$K${sorted_students_end},MATCH(Y{row_index},$M$2:$M${sorted_students_end},0)),""),"")'
            )

    def _render_presentation_sheet(self, worksheet, *, view_model) -> None:
        worksheet["A1"] = view_model.title
        worksheet["A1"].font = _TITLE_FONT
        worksheet["A2"] = view_model.class_name
        worksheet["A2"].font = Font(size=12, bold=True)
        worksheet["A3"] = view_model.generated_label
        worksheet["A3"].font = _META_FONT
        worksheet.column_dimensions["A"].width = 10
        worksheet.column_dimensions["B"].width = 30
        worksheet.column_dimensions[_PRESENTATION_ROW_TYPE_COLUMN].hidden = True
        worksheet.column_dimensions[_PRESENTATION_GROUP_COLUMN].hidden = True
        worksheet.column_dimensions[_PRESENTATION_NUMBER_COLUMN].hidden = True
        worksheet.column_dimensions[_PRESENTATION_STUDENT_COLUMN].hidden = True

        presentation_end_row = _PRESENTATION_START_ROW + view_model.presentation_row_capacity - 1
        for sheet_row in range(_PRESENTATION_START_ROW, presentation_end_row + 1):
            helper_row = (sheet_row - _PRESENTATION_START_ROW) + 2
            worksheet[f"{_PRESENTATION_ROW_TYPE_COLUMN}{sheet_row}"] = (
                f"='{_HELPER_SHEET_TITLE}'!Z{helper_row}"
            )
            worksheet[f"{_PRESENTATION_GROUP_COLUMN}{sheet_row}"] = (
                f"='{_HELPER_SHEET_TITLE}'!AA{helper_row}"
            )
            worksheet[f"{_PRESENTATION_NUMBER_COLUMN}{sheet_row}"] = (
                f"='{_HELPER_SHEET_TITLE}'!AB{helper_row}"
            )
            worksheet[f"{_PRESENTATION_STUDENT_COLUMN}{sheet_row}"] = (
                f"='{_HELPER_SHEET_TITLE}'!AC{helper_row}"
            )
            worksheet[f"A{sheet_row}"] = (
                f'=IF(${_PRESENTATION_ROW_TYPE_COLUMN}{sheet_row}="header","Nr",'
                f'IF(${_PRESENTATION_ROW_TYPE_COLUMN}{sheet_row}="member",'
                f'${_PRESENTATION_NUMBER_COLUMN}{sheet_row},""))'
            )
            worksheet[f"B{sheet_row}"] = (
                f'=IF(${_PRESENTATION_ROW_TYPE_COLUMN}{sheet_row}="heading",'
                f"${_PRESENTATION_GROUP_COLUMN}{sheet_row},"
                f'IF(${_PRESENTATION_ROW_TYPE_COLUMN}{sheet_row}="header","Elev",'
                f'IF(${_PRESENTATION_ROW_TYPE_COLUMN}{sheet_row}="member",'
                f'${_PRESENTATION_STUDENT_COLUMN}{sheet_row},"")))'
            )
            for column in ("A", "B"):
                worksheet[f"{column}{sheet_row}"].alignment = Alignment(
                    vertical="center",
                    horizontal="left",
                    wrap_text=True,
                )

        worksheet.conditional_formatting.add(
            f"A{_PRESENTATION_START_ROW}:B{presentation_end_row}",
            FormulaRule(
                formula=[f'${_PRESENTATION_ROW_TYPE_COLUMN}{_PRESENTATION_START_ROW}="heading"'],
                border=_PRESENTATION_BORDER,
                fill=_SECTION_FILL,
                font=_SECTION_FONT,
            ),
        )
        worksheet.conditional_formatting.add(
            f"A{_PRESENTATION_START_ROW}:B{presentation_end_row}",
            FormulaRule(
                formula=[f'${_PRESENTATION_ROW_TYPE_COLUMN}{_PRESENTATION_START_ROW}="header"'],
                border=_PRESENTATION_BORDER,
                fill=_HEADER_FILL,
                font=_HEADER_FONT,
            ),
        )
        worksheet.conditional_formatting.add(
            f"A{_PRESENTATION_START_ROW}:B{presentation_end_row}",
            FormulaRule(
                formula=[f'${_PRESENTATION_ROW_TYPE_COLUMN}{_PRESENTATION_START_ROW}="member"'],
                border=_PRESENTATION_BORDER,
            ),
        )

        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_title_rows = "1:3"
        worksheet.print_area = f"A1:B{presentation_end_row}"

    def _write_helper_headers(self, worksheet) -> None:
        headers = {
            "A1": "source_group",
            "B1": "source_group_order",
            "C1": "source_member_order",
            "D1": "source_student_name",
            "E1": "source_assigned",
            "F1": "source_sort_rank",
            "H1": "sorted_group",
            "I1": "sorted_group_order",
            "J1": "sorted_member_order",
            "K1": "sorted_student_name",
            "L1": "group_member_index",
            "M1": "member_presentation_row",
            "N1": "registry_group",
            "O1": "registry_group_order",
            "P1": "registry_sort_rank",
            "R1": "sorted_registry_group",
            "S1": "sorted_registry_order",
            "T1": "registry_member_count",
            "U1": "heading_presentation_row",
            "V1": "header_presentation_row",
            "W1": "registry_lookup_key",
            "Y1": "presentation_slot",
            "Z1": "presentation_row_type",
            "AA1": "presentation_group",
            "AB1": "presentation_number",
            "AC1": "presentation_student_name",
        }
        for cell_reference, value in headers.items():
            worksheet[cell_reference] = value
        for column in range(1, 30):
            worksheet.column_dimensions[get_column_letter(column)].hidden = True
