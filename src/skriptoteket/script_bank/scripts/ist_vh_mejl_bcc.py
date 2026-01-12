"""
Extrahera vårdnadshavares e-postadresser från en klasslista (CSV/XLSX).

Vanligt användningsområden:
- Du har en export från IST (eller liknande system) med elevrader.
- I exporten finns en eller flera kolumner med vårdnadshavares e‑postadresser.
- Du vill snabbt få ut en semikolonseparerad lista som går att klistra in i Outlooks BCC-fält.

Runner-kontrakt:
- Entrypoint: run_tool(input_dir: str, output_dir: str) -> dict
- Input: En eller flera CSV/XLSX-filer med rubrikrad (multi-fil via SKRIPTOTEKET_INPUT_MANIFEST)
- Output: Contract v2 dict with typed outputs + en fil `emails_<timestamp>.txt` som artifact
"""

from __future__ import annotations

import json
import os
import re
from collections.abc import Iterable
from datetime import datetime, timezone
from pathlib import Path
from typing import TypedDict

EMAIL_RE = re.compile(r"([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,15})")


class ManifestFile(TypedDict):
    name: str
    path: str
    bytes: int


class FileExtractionResult(TypedDict):
    rows: int
    emails: list[str]
    duplicates_removed: int
    column_email_counts: dict[str, int]
    error: str | None


SUPPORTED_INPUT_SUFFIXES = {".xlsx", ".csv", ".txt"}

# Ledtrådar som ofta finns i kolumnrubriker (svenska + engelska varianter)
GUARDIAN_HINTS = [
    "vardnad",
    "vårdnad",
    "foralder",
    "förälder",
    "parent",
    "guardian",
    "guardians",
]
EMAIL_HINTS = [
    "e-post",
    "epost",
    "email",
    "e-mail",
    "mejl",
    "mail",
]


def _read_input_manifest_files() -> list[ManifestFile]:
    raw = os.environ.get("SKRIPTOTEKET_INPUT_MANIFEST", "")
    if not raw.strip():
        return []
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return []
    if not isinstance(payload, dict):
        return []
    files = payload.get("files")
    if not isinstance(files, list):
        return []

    result: list[ManifestFile] = []
    for item in files:
        if not isinstance(item, dict):
            continue
        name = item.get("name")
        path = item.get("path")
        bytes_ = item.get("bytes")
        if isinstance(name, str) and isinstance(path, str) and isinstance(bytes_, int):
            result.append({"name": name, "path": path, "bytes": bytes_})
    return result


def _select_input_files(*, input_dir: Path) -> list[ManifestFile]:
    manifest_files: list[ManifestFile] = _read_input_manifest_files()
    if manifest_files:
        supported = [
            file
            for file in manifest_files
            if Path(file["name"]).suffix.lower() in SUPPORTED_INPUT_SUFFIXES
        ]
        return supported if supported else manifest_files

    if not input_dir.is_dir():
        return []

    scanned: list[ManifestFile] = [
        {"name": path.name, "path": str(path), "bytes": path.stat().st_size}
        for path in sorted(input_dir.glob("*"))
        if path.is_file()
    ]
    supported = [
        file for file in scanned if Path(file["name"]).suffix.lower() in SUPPORTED_INPUT_SUFFIXES
    ]
    return supported if supported else scanned


def _column_matches_hints(col: str, hints: list[str]) -> bool:
    col_lower = col.lower()
    return any(hint in col_lower for hint in hints)


def prioritized_columns(columns: list[str]) -> list[str]:
    """Returnera kolumner i prioriterad ordning för mejl-extrahering.

    Prioritet:
    1) Kolumner som matchar både vårdnadshavare- och mejl-ledtrådar
    2) Kolumner som matchar mejl-ledtrådar
    3) Kolumner som matchar vårdnadshavare-ledtrådar
    4) Alla andra kolumner (fallback)
    """
    guardian_email = [
        c
        for c in columns
        if _column_matches_hints(c, GUARDIAN_HINTS) and _column_matches_hints(c, EMAIL_HINTS)
    ]
    email_only = [
        c for c in columns if _column_matches_hints(c, EMAIL_HINTS) and c not in guardian_email
    ]
    guardian_only = [
        c for c in columns if _column_matches_hints(c, GUARDIAN_HINTS) and c not in guardian_email
    ]
    other = [c for c in columns if c not in guardian_email + email_only + guardian_only]

    return guardian_email + email_only + guardian_only + other


def harvest_emails_from_cells(cells: Iterable[str]) -> list[str]:
    """Plocka ut unika mejladresser från celler (normaliseras till lowercase)."""
    seen: set[str] = set()
    result: list[str] = []
    for cell in cells:
        if not cell:
            continue
        for match in EMAIL_RE.findall(str(cell)):
            email = match.lower()
            if email not in seen:
                seen.add(email)
                result.append(email)
    return result


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    """Läs CSV med enkel avgränsardetektering (komma, semikolon, tab)."""
    import csv

    content = path.read_text(encoding="utf-8-sig")

    for delimiter in [",", ";", "\t"]:
        try:
            reader = csv.reader(content.splitlines(), delimiter=delimiter)
            rows = list(reader)
            if rows and len(rows[0]) > 1:
                return rows[0], rows[1:]
        except csv.Error:
            continue

    reader = csv.reader(content.splitlines())
    rows = list(reader)
    return rows[0] if rows else [], rows[1:] if len(rows) > 1 else []


def read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    """Läs XLSX med openpyxl (första bladet)."""
    import warnings

    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        wb.close()
        return [], []

    rows = [[str(cell.value or "") for cell in row] for row in ws.iter_rows()]
    wb.close()

    if not rows:
        return [], []
    return rows[0], rows[1:]


def _extract_emails_from_file(*, path: Path) -> FileExtractionResult:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        headers, data_rows = read_xlsx(path)
    elif suffix in {".csv", ".txt"}:
        headers, data_rows = read_csv(path)
    else:
        return {
            "rows": 0,
            "emails": [],
            "duplicates_removed": 0,
            "column_email_counts": {},
            "error": f"Filtypen '{suffix}' stöds inte. Använd .csv eller .xlsx.",
        }

    if not headers:
        return {
            "rows": 0,
            "emails": [],
            "duplicates_removed": 0,
            "column_email_counts": {},
            "error": "Filen verkar tom eller saknar kolumnrubriker.",
        }

    ordered_cols = prioritized_columns(headers)
    col_indices = {col: headers.index(col) for col in ordered_cols}

    all_emails: list[str] = []
    total_email_occurrences = 0
    col_email_counts: dict[str, int] = {}

    for col in ordered_cols:
        idx = col_indices[col]
        cells = [row[idx] for row in data_rows if idx < len(row)]
        emails = harvest_emails_from_cells(cells)
        total_email_occurrences += len(emails)
        new_emails = [e for e in emails if e not in all_emails]
        col_email_counts[col] = len(new_emails)
        all_emails.extend(new_emails)

    duplicates_removed = total_email_occurrences - len(all_emails)
    return {
        "rows": len(data_rows),
        "emails": all_emails,
        "duplicates_removed": duplicates_removed,
        "column_email_counts": col_email_counts,
        "error": None,
    }


def run_tool(input_dir: str, output_dir: str) -> dict:
    """Entrypoint: extrahera vårdnadshavares mejl från en eller flera klasslistor."""
    input_root = Path(input_dir)
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)

    if not input_root.is_dir():
        return {
            "outputs": [
                {
                    "kind": "notice",
                    "level": "error",
                    "message": "Indatakatalogen kunde inte hittas eller läsas.",
                }
            ],
            "next_actions": [],
            "state": None,
        }

    input_files = _select_input_files(input_dir=input_root)

    seen_global: set[str] = set()
    global_emails: list[str] = []
    total_rows = 0
    total_duplicates_in_files = 0
    total_skipped_already_seen = 0

    per_file_rows: list[dict[str, object]] = []
    aggregated_column_counts: dict[str, int] = {}

    for file in input_files:
        file_path = Path(file["path"])
        try:
            result = _extract_emails_from_file(path=file_path)
        except Exception as exc:  # noqa: BLE001 - tool script robustness
            result = {
                "rows": 0,
                "emails": [],
                "duplicates_removed": 0,
                "column_email_counts": {},
                "error": f"Kunde inte läsa filen: {exc}",
            }

        if result["error"] is not None:
            per_file_rows.append(
                {
                    "file": file["name"],
                    "rows": 0,
                    "unique_emails": 0,
                    "new_emails": 0,
                    "duplicates_in_file": 0,
                    "skipped_already_seen": 0,
                    "status": "error",
                    "message": result["error"],
                }
            )
            continue

        file_emails = result["emails"]
        new_emails = [email for email in file_emails if email not in seen_global]
        skipped = len(file_emails) - len(new_emails)
        for email in new_emails:
            seen_global.add(email)
        global_emails.extend(new_emails)

        total_rows += result["rows"]
        total_duplicates_in_files += result["duplicates_removed"]
        total_skipped_already_seen += skipped

        for column, count in result["column_email_counts"].items():
            aggregated_column_counts[column] = aggregated_column_counts.get(column, 0) + count

        per_file_rows.append(
            {
                "file": file["name"],
                "rows": result["rows"],
                "unique_emails": len(file_emails),
                "new_emails": len(new_emails),
                "duplicates_in_file": result["duplicates_removed"],
                "skipped_already_seen": skipped,
                "status": "ok",
                "message": "",
            }
        )

    if not per_file_rows or not any(row["status"] == "ok" for row in per_file_rows):
        if len(per_file_rows) == 1 and per_file_rows[0]["status"] == "error":
            return {
                "outputs": [
                    {
                        "kind": "notice",
                        "level": "error",
                        "message": str(per_file_rows[0]["message"]),
                    }
                ],
                "next_actions": [],
                "state": None,
            }
        return {
            "outputs": [
                {
                    "kind": "notice",
                    "level": "error",
                    "message": "Inga giltiga filer kunde behandlas. Ladda upp .csv eller .xlsx.",
                },
                {
                    "kind": "table",
                    "title": "Filer",
                    "columns": [
                        {"key": "file", "label": "Fil"},
                        {"key": "status", "label": "Status"},
                        {"key": "message", "label": "Meddelande"},
                    ],
                    "rows": [
                        {"file": row["file"], "status": row["status"], "message": row["message"]}
                        for row in per_file_rows
                    ],
                },
            ],
            "next_actions": [],
            "state": None,
        }

    if not global_emails:
        return {
            "outputs": [
                {
                    "kind": "notice",
                    "level": "warning",
                    "message": "Inga e-postadresser hittades i de uppladdade filerna.",
                },
                {
                    "kind": "table",
                    "title": "Filer",
                    "columns": [
                        {"key": "file", "label": "Fil"},
                        {"key": "rows", "label": "Rader"},
                        {"key": "unique_emails", "label": "E-post (unika)"},
                        {"key": "status", "label": "Status"},
                        {"key": "message", "label": "Meddelande"},
                    ],
                    "rows": [
                        {
                            "file": row["file"],
                            "rows": row["rows"],
                            "unique_emails": row["unique_emails"],
                            "status": row["status"],
                            "message": row["message"],
                        }
                        for row in per_file_rows
                    ],
                },
            ],
            "next_actions": [],
            "state": None,
        }

    email_string = ";".join(global_emails)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    artifact_name = f"emails_{timestamp}.txt"
    artifact_path = output / artifact_name
    artifact_path.write_text(email_string, encoding="utf-8")

    stats_table_rows = [
        {"column": column, "count": count}
        for column, count in sorted(
            aggregated_column_counts.items(), key=lambda item: item[1], reverse=True
        )
        if count > 0
    ]

    dup_parts: list[str] = []
    if total_duplicates_in_files > 0:
        dup_parts.append(f"{total_duplicates_in_files} dubbletter i filer filtrerades bort")
    if total_skipped_already_seen > 0:
        dup_parts.append(
            f"{total_skipped_already_seen} e-postadresser fanns redan från tidigare filer"
        )
    dup_info = f" ({'; '.join(dup_parts)})" if dup_parts else ""

    preview_count = min(10, len(global_emails))
    preview = ", ".join(global_emails[:preview_count])
    remaining = len(global_emails) - preview_count
    preview_suffix = f" ... (+{remaining} till)" if remaining > 0 else ""
    ok_files_count = sum(1 for row in per_file_rows if row["status"] == "ok")

    return {
        "outputs": [
            {
                "kind": "notice",
                "level": "info",
                "message": (
                    f"{len(global_emails)} unika e-postadresser extraherades "
                    f"från {total_rows} rader i {ok_files_count} filer."
                    f"{dup_info}"
                ),
            },
            {
                "kind": "markdown",
                "markdown": (
                    f"Ladda ner **{artifact_name}** och klistra in innehållet i Outlooks BCC-fält "
                    "(semikolonseparerat).\n\n"
                    f"**Förhandsvisning:** {preview}{preview_suffix}"
                ),
            },
            {
                "kind": "table",
                "title": "Filer",
                "columns": [
                    {"key": "file", "label": "Fil"},
                    {"key": "rows", "label": "Rader"},
                    {"key": "unique_emails", "label": "E-post (unika)"},
                    {"key": "new_emails", "label": "Nya (sammanlagt)"},
                    {"key": "duplicates_in_file", "label": "Dubbletter (i fil)"},
                    {"key": "skipped_already_seen", "label": "Redan sedda"},
                    {"key": "status", "label": "Status"},
                    {"key": "message", "label": "Meddelande"},
                ],
                "rows": per_file_rows,
            },
            {
                "kind": "table",
                "title": "Statistik per kolumn (sammanlagt)",
                "columns": [
                    {"key": "column", "label": "Kolumn"},
                    {"key": "count", "label": "Nya e-postadresser"},
                ],
                "rows": stats_table_rows,
            },
        ],
        "next_actions": [],
        "state": None,
    }


if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <input_dir> <output_dir>")
        raise SystemExit(1)

    result = run_tool(sys.argv[1], sys.argv[2])
    print(json.dumps(result, indent=2, ensure_ascii=False))
