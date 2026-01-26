"""
Gruppgeneratorn: skapa elevgrupper från klasslistor (CSV/XLSX).

- Stöd för sparade klasser via settings (memory.json)
- Valfri hänsyn till tidigare grupper (minska upprepade par)
"""

from __future__ import annotations

import csv
import json
import random
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Literal, NotRequired, TypedDict

from skriptoteket_toolkit import (
    get_action_parts,
    list_input_files,
    read_inputs,
    read_settings,
)

SUPPORTED_ROSTER_SUFFIXES = {".csv", ".xlsx"}
DEFAULT_GROUP_SIZE = 3
MAX_SHUFFLES = 200


class ToolResult(TypedDict):
    outputs: list[dict[str, object]]
    next_actions: list[dict[str, object]]
    state_update: dict[str, object]
    promotions: NotRequired[dict[str, object]]


def _notice(level: Literal["info", "warning", "error"], message: str) -> dict[str, object]:
    return {"kind": "notice", "level": level, "message": message}


def _markdown(content: str) -> dict[str, object]:
    return {"kind": "markdown", "markdown": content}


def _table(title: str, rows: list[dict[str, object]]) -> dict[str, object]:
    return {
        "kind": "table",
        "title": title,
        "columns": [
            {"key": "group", "label": "Grupp"},
            {"key": "members", "label": "Elever"},
        ],
        "rows": rows,
    }


def _normalize_name(value: str) -> str:
    return " ".join(value.strip().split())


def _dedupe_names(names: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for name in names:
        cleaned = _normalize_name(name)
        if not cleaned:
            continue
        key = cleaned.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def _detect_name_column(headers: list[str]) -> int:
    for idx, header in enumerate(headers):
        candidate = header.strip().casefold()
        if "namn" in candidate or "name" in candidate:
            return idx
    return 0


def _read_csv_rows(path: Path) -> list[list[str]]:
    content = path.read_text(encoding="utf-8-sig")
    # Try sniffing, fallback to common delimiters
    for delimiter in [",", ";", "\t"]:
        try:
            reader = csv.reader(content.splitlines(), delimiter=delimiter)
            rows = list(reader)
        except csv.Error:
            continue
        if rows and len(rows[0]) > 1:
            return rows
    # Fallback default
    reader = csv.reader(content.splitlines())
    return list(reader)


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return []
        rows = [[str(cell.value or "").strip() for cell in row] for row in ws.iter_rows()]
        wb.close()
        return rows
    except Exception:
        return []


def _extract_names_from_rows(rows: list[list[str]]) -> list[str]:
    if not rows:
        return []
    headers = rows[0]
    column_index = _detect_name_column(headers)
    names = []
    for row in rows[1:]:
        if column_index >= len(row):
            continue
        names.append(row[column_index])
    return _dedupe_names(names)


def _parse_roster_file(path: Path) -> tuple[list[str], str | None]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_ROSTER_SUFFIXES:
        return ([], f"Filtypen '{suffix}' stöds inte. Använd .csv eller .xlsx.")

    if suffix == ".xlsx":
        rows = _read_xlsx_rows(path)
        if not rows:
            return ([], "Kunde inte läsa Excel-filen (openpyxl saknas eller filen är tom/trasig).")
        return (_extract_names_from_rows(rows), None)

    rows = _read_csv_rows(path)
    if not rows:
        return ([], "CSV-filen verkar vara tom eller oläslig.")
    return (_extract_names_from_rows(rows), None)


def _parse_saved_classes(settings: dict[str, object]) -> tuple[dict[str, list[str]], str | None]:
    raw = settings.get("saved_classes_json", "")

    # Also look for session-promoted classes
    session_classes: dict[str, list[str]] = {}
    for item in list_input_files():
        if item["name"] == "session_classes.json":
            try:
                data = json.loads(Path(item["path"]).read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    for k, v in data.items():
                        if isinstance(v, list):
                            session_classes[str(k)] = [str(i) for i in v]
            except Exception:
                pass

    if raw is None:
        return (session_classes, None)

    payload: dict[str, object] = {}
    if isinstance(raw, dict):
        payload = raw
    elif isinstance(raw, str):
        if not raw.strip():
            return (session_classes, None)
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            return (session_classes, "Sparade klasser (JSON) i settings innehåller ogiltig JSON.")
    else:
        return (
            session_classes,
            "Sparade klasser (JSON) i settings måste vara text eller ett JSON-objekt.",
        )

    if not isinstance(payload, dict):
        return (session_classes, "Sparade klasser (JSON) i settings måste vara ett objekt.")

    classes: dict[str, list[str]] = dict(session_classes)
    for key, value in payload.items():
        if not isinstance(key, str):
            continue
        if isinstance(value, list):
            names = [str(item) for item in value if str(item).strip()]
            classes[key.strip()] = _dedupe_names(names)
    return (classes, None)


def _parse_previous_groups(raw: str) -> tuple[list[list[str]], str | None]:
    if not raw.strip():
        return ([], None)

    # Try JSON first
    try:
        payload = json.loads(raw)
        if isinstance(payload, list):
            groups = []
            for item in payload:
                if isinstance(item, list):
                    clean_group = [str(entry).strip() for entry in item if str(entry).strip()]
                    if clean_group:
                        groups.append(clean_group)
            return (groups, None)
    except json.JSONDecodeError:
        pass

    # Fallback: simple text parsing (lines with comma/semicolon)
    groups: list[list[str]] = []
    for line in raw.splitlines():
        if not line.strip():
            continue
        parts = [p.strip() for p in re.split(r"[;,]", line) if p.strip()]
        if parts:
            groups.append(parts)

    if not groups:
        return ([], "Kunde inte tolka tidigare grupper (varken som JSON eller text).")
    return (groups, None)


def _build_action_fields(saved_classes: dict[str, list[str]]) -> list[dict[str, object]]:
    fields: list[dict[str, object]] = []
    if saved_classes:
        options = [
            {"value": key, "label": key} for key in sorted(saved_classes.keys(), key=str.casefold)
        ]
        fields.append(
            {
                "name": "saved_class",
                "label": "Sparad klass (valfri)",
                "kind": "enum",
                "options": options,
            }
        )
    fields.extend(
        [
            {
                "name": "group_size",
                "label": "Gruppstorlek",
                "kind": "integer",
            },
            {
                "name": "class_name",
                "label": "Klassnamn (för sparad klass eller för att spara ny)",
                "kind": "string",
            },
            {
                "name": "group_set_name",
                "label": "Namn på gruppindelningen",
                "kind": "string",
            },
            {
                "name": "previous_groups_file",
                "label": "Välj tidigare grupper (från session/valv)",
                "kind": "file_ref",
                "min": 0,
                "max": 1,
            },
            {
                "name": "previous_groups",
                "label": "Klistra in tidigare grupper (valfritt om fil ej väljs)",
                "kind": "text",
            },
        ]
    )
    return fields


def _build_actions(saved_classes: dict[str, list[str]]) -> list[dict[str, object]]:
    return [
        {
            "action_id": "generate",
            "label": "Skapa grupper",
            "kind": "form",
            "fields": _build_action_fields(saved_classes),
        }
    ]


def _pairs_for_groups(groups: list[list[str]]) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for group in groups:
        cleaned = _dedupe_names(group)
        for idx, name in enumerate(cleaned):
            for other in cleaned[idx + 1 :]:
                a, b = sorted((name.casefold(), other.casefold()))
                pairs.add((a, b))
    return pairs


def _score_groups(groups: list[list[str]], previous_pairs: set[tuple[str, str]]) -> int:
    if not previous_pairs:
        return 0
    score = 0
    for group in groups:
        cleaned = _dedupe_names(group)
        for idx, name in enumerate(cleaned):
            for other in cleaned[idx + 1 :]:
                a, b = sorted((name.casefold(), other.casefold()))
                if (a, b) in previous_pairs:
                    score += 1
    return score


def _chunk_groups(students: list[str], group_size: int) -> list[list[str]]:
    return [students[i : i + group_size] for i in range(0, len(students), group_size)]


def _build_groups(
    students: list[str],
    group_size: int,
    previous_pairs: set[tuple[str, str]],
) -> tuple[list[list[str]], int]:
    if not students:
        return ([], 0)
    rng = random.Random()
    best_groups: list[list[str]] = []
    best_score = float("inf")
    attempts = MAX_SHUFFLES if previous_pairs else 1
    pool = students[:]

    for _ in range(attempts):
        rng.shuffle(pool)
        groups = _chunk_groups(pool, group_size)
        score = _score_groups(groups, previous_pairs)
        if score < best_score:
            best_score = float(score)
            best_groups = [group[:] for group in groups]
            if score == 0:
                break
    return (best_groups, int(best_score if best_score != float("inf") else 0))


def _select_roster_file(input_dir: Path) -> Path | None:
    manifest_files = list_input_files()
    for item in manifest_files:
        path = Path(str(item["path"]))
        if path.suffix.lower() in SUPPORTED_ROSTER_SUFFIXES:
            return path

    if not input_dir.is_dir():
        return None
    for path in sorted(input_dir.glob("*")):
        if path.suffix.lower() in SUPPORTED_ROSTER_SUFFIXES and path.is_file():
            return path
    return None


def run_tool(input_dir: str, output_dir: str) -> ToolResult:
    inputs_raw = read_inputs()
    settings_raw = read_settings()
    action_id, action_input_raw, _ = get_action_parts()

    inputs: dict[str, object] = inputs_raw if isinstance(inputs_raw, dict) else {}
    settings: dict[str, object] = settings_raw if isinstance(settings_raw, dict) else {}
    action_input: dict[str, object] = action_input_raw if isinstance(action_input_raw, dict) else {}

    payload = action_input if action_id == "generate" else inputs

    group_size = int(payload.get("group_size") or settings.get("default_group_size") or 0)
    if group_size < 2:
        group_size = DEFAULT_GROUP_SIZE

    class_name = str(payload.get("class_name") or "").strip()
    group_set_name = str(payload.get("group_set_name") or "").strip()
    previous_groups_raw = str(payload.get("previous_groups") or "")

    # Strategy: Read previous groups from file if provided, else fallback to text field
    previous_groups_file_refs = payload.get("previous_groups_file")
    if isinstance(previous_groups_file_refs, list) and previous_groups_file_refs:
        # Take the first one (usually max=1 anyway)
        file_ref = previous_groups_file_refs[0]
        # In Contract V3, selected files are staged into input_dir
        # We find it by matching name in list_input_files
        for item in list_input_files():
            if item.get("ref") == file_ref:
                try:
                    previous_groups_raw = Path(item["path"]).read_text(encoding="utf-8")
                    break
                except Exception:
                    pass

    saved_classes, saved_error = _parse_saved_classes(settings)
    if saved_error:
        saved_classes = {}

    roster_path = _select_roster_file(Path(input_dir))

    # --- Initial State (No Action) ---
    if action_id is None:
        outputs: list[dict[str, object]] = []
        if saved_error:
            outputs.append(_notice("warning", saved_error))

        has_saved = bool(saved_classes)

        if roster_path is not None:
            students, roster_error = _parse_roster_file(roster_path)
            if roster_error:
                outputs.append(_notice("error", roster_error))
            else:
                outputs.append(
                    _notice("info", f"Hittade {len(students)} elever i '{roster_path.name}'.")
                )
                if students:
                    preview = ", ".join(students[:10])
                    if len(students) > 10:
                        preview += "..."
                    outputs.append(_markdown(f"**Exempel på namn:** {preview}"))
        elif has_saved:
            outputs.append(
                _notice("info", "Inga filer uppladdade. Du kan välja en sparad klass nedan.")
            )
        else:
            outputs.append(
                _notice(
                    "error",
                    "Ingen klasslista hittades. Ladda upp en fil eller spara klasser i settings.",
                )
            )

        outputs.append(_markdown("Fyll i formuläret nedan och klicka **Skapa grupper**."))

        return {
            "outputs": outputs,
            "next_actions": _build_actions(saved_classes),
            "state_update": {"kind": "no_change"},
        }

    # --- Handle Action ---
    if action_id != "generate":
        return {
            "outputs": [_notice("error", f"Okänd action_id: '{action_id}'.")],
            "next_actions": _build_actions(saved_classes),
            "state_update": {"kind": "no_change"},
        }

    selected_saved = str(payload.get("saved_class") or "").strip()
    students: list[str] = []
    roster_error: str | None = None
    source_label = ""

    # Strategy: 1. Selected Saved, 2. Matching Class Name (Saved), 3. Uploaded File
    if selected_saved and selected_saved in saved_classes:
        students = saved_classes[selected_saved]
        class_name = selected_saved
        source_label = f"Sparad klass: {selected_saved}"
    elif class_name and class_name in saved_classes:
        students = saved_classes[class_name]
        source_label = f"Sparad klass: {class_name}"
    elif roster_path is not None:
        students, roster_error = _parse_roster_file(roster_path)
        if not class_name:
            class_name = roster_path.stem
        source_label = f"Uppladdad fil: {roster_path.name}"
    elif class_name:
        roster_error = f"Hittade ingen sparad klass med namnet '{class_name}'."
    else:
        roster_error = (
            "Ingen klasslista hittades. Ladda upp en fil eller ange ett sparat klassnamn."
        )

    if roster_error or not students:
        return {
            "outputs": [
                _notice("error", roster_error or "Inga elever hittades i klasslistan."),
                _markdown(
                    "Tips: Lägg klasslistor i settings (memory.json) för att återanvända dem."
                ),
            ],
            "next_actions": _build_actions(saved_classes),
            "state_update": {"kind": "no_change"},
        }

    if group_size > len(students):
        return {
            "outputs": [
                _notice(
                    "error",
                    f"Gruppstorlek ({group_size}) är större än antalet elever ({len(students)}).",
                )
            ],
            "next_actions": _build_actions(saved_classes),
            "state_update": {"kind": "no_change"},
        }

    previous_groups, previous_error = _parse_previous_groups(previous_groups_raw)
    previous_pairs = _pairs_for_groups(previous_groups)

    groups, repeats = _build_groups(students, group_size, previous_pairs)

    group_label = group_set_name or "Grupp"
    rows: list[dict[str, object]] = []
    markdown_lines = [f"**{group_set_name or 'Gruppindelning'}**\n"]
    for idx, group in enumerate(groups, start=1):
        name = f"{group_label} {idx}".strip()
        members = ", ".join(group)
        rows.append({"group": name, "members": members})
        markdown_lines.append(f"{idx}. {members}")

    # Create artifacts and promotions
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    # Artifact 1: Group list (TXT)
    artifact_name = f"grupper_{timestamp}.txt"
    (output_path / artifact_name).write_text("\n".join(markdown_lines), encoding="utf-8")

    promotions = {}
    # If we have new/updated classes, promote them to the session
    if class_name and (roster_path is not None or students):
        updated_classes = dict(saved_classes)
        updated_classes[class_name] = students

        # Write to a hidden artifact for session promotion
        session_file = "session_classes.json"
        (output_path / session_file).write_text(json.dumps(updated_classes), encoding="utf-8")

        promotions = {
            "requests": [
                {
                    "kind": "session",
                    "name": session_file,
                    "source_path": f"output/{session_file}",
                }
            ]
        }

    outputs: list[dict[str, object]] = [
        _notice(
            "info", f"Elever: {len(students)} | Gruppstorlek: {group_size} | Grupper: {len(groups)}"
        ),
        _notice("info", source_label),
        _markdown("\n".join(markdown_lines)),
        _table("Gruppindelning", rows),
        _notice("info", f"Artefakt skapad: {artifact_name}"),
    ]

    if saved_error:
        outputs.append(_notice("warning", saved_error))
    if previous_error:
        outputs.append(_notice("warning", previous_error))
    if previous_pairs:
        outputs.append(_notice("info", f"Upprepade par mot tidigare grupper: {repeats}"))

    # Offer to save class if from file (for PERMANENT storage in Settings)
    if class_name and roster_path is not None:
        updated_classes = dict(saved_classes)
        updated_classes[class_name] = students
        outputs.append(
            _markdown(
                "Klassen har sparats **temporärt** för denna session. "
                "För att spara den **permanent**, kopiera JSON-blocket nedan till "
                "verktygets **Settings**."
            )
        )
        outputs.append(
            {
                "kind": "json",
                "title": "Sparade klasser (JSON)",
                "value": updated_classes,
            }
        )

    if saved_classes:
        outputs.append(
            _markdown("Tillgängliga sparade klasser: " + ", ".join(sorted(saved_classes.keys())))
        )

    result: ToolResult = {
        "outputs": outputs,
        "next_actions": _build_actions(saved_classes),
        "state_update": {"kind": "no_change"},
    }
    if promotions:
        result["promotions"] = promotions

    return result
