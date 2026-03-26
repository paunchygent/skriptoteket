"""Unit tests for the classroom-planner grouping XLSX renderer.

Purpose:
    Guard the teacher-facing workbook shape so the grouping export stays
    editable while keeping the linked presentation sheet stable enough to
    share or print outside the app.

Relationships:
    - Exercises `GroupingXlsxRenderer`.
    - Builds the workbook through the grouping XLSX workbook view model.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from skriptoteket.application.curated_apps.classroom_planner.exports import (
    GroupingExportPresentation,
    GroupingPresentationGroup,
    GroupingPresentationMember,
    build_grouping_xlsx_view_model,
)
from skriptoteket.infrastructure.curated_apps.apps.classroom_planner.grouping_xlsx_renderer import (
    GroupingXlsxRenderer,
)


def _presentation() -> GroupingExportPresentation:
    return GroupingExportPresentation(
        draft_id=uuid4(),
        class_name="Klass 7A",
        title="Gruppindelning",
        filename_stem="klass-7a-gruppindelning",
        groups=(
            GroupingPresentationGroup(
                group_label="Grupp 1",
                group_order=0,
                members=(
                    GroupingPresentationMember(member_order=1, display_name="Ada Lovelace"),
                    GroupingPresentationMember(member_order=2, display_name="Bo Berg"),
                ),
            ),
            GroupingPresentationGroup(
                group_label="Grupp 2",
                group_order=1,
                members=(
                    GroupingPresentationMember(
                        member_order=1,
                        display_name="Linus Torvalds",
                    ),
                ),
            ),
        ),
    )


def _render_workbook():
    view_model = build_grouping_xlsx_view_model(
        presentation=_presentation(),
        generated_at=datetime(2026, 3, 26, 12, 34, tzinfo=timezone.utc),
        unassigned_student_names=("Grace Hopper",),
    )
    workbook_bytes = GroupingXlsxRenderer().render(view_model=view_model)
    return load_workbook(BytesIO(workbook_bytes))


@pytest.mark.unit
def test_renderer_uses_two_visible_teacher_sheets_and_hidden_helper_sheet():
    workbook = _render_workbook()

    assert workbook.sheetnames == ["Redigera grupper", "Dela och exportera", "_PresentationData"]
    assert workbook.active.title == "Dela och exportera"
    assert workbook["_PresentationData"].sheet_state == "hidden"


@pytest.mark.unit
def test_renderer_writes_one_edit_table_with_locked_headers_and_unassigned_rows():
    workbook = _render_workbook()
    sheet = workbook["Redigera grupper"]

    assert sheet["A1"].value == "Nr i grupp"
    assert sheet["B1"].value == "Elev"
    assert sheet["C1"].value == "Grupp (välj)"
    assert sheet["F1"].value == "Grupp"
    assert sheet["G1"].value == "Gruppordning (välj)"
    assert sheet.freeze_panes == "A2"
    assert list(sheet.tables) == ["tblRedigeraGrupper", "tblGruppregister"]
    assert sheet["A2"].value == 1
    assert sheet["B2"].value == "Ada Lovelace"
    assert sheet["C2"].value == "Grupp 1"
    assert sheet["A5"].value == 1
    assert sheet["B5"].value == "Grace Hopper"
    assert sheet["C5"].value is None
    assert sheet["F2"].value == "Grupp 1"
    assert sheet["G2"].value == 1
    assert sheet["F3"].value == "Grupp 2"
    assert sheet["G3"].value == 2
    validations = list(sheet.data_validations.dataValidation)
    assert len(validations) == 2
    assert validations[0].formula1 == "=$F$2:$F$3"
    assert str(validations[0].sqref) == "C2:C5"
    assert validations[1].formula1 == '"1,2"'
    assert str(validations[1].sqref) == "G2:G3"
    assert sheet["I1"].value == "Ändra bara detta"
    assert (
        sheet["J4"].value
        == "Ändra bara gruppordningen i kolumn G. Exportera en ny fil för större ändringar."
    )
    assert sheet["J5"].value == "Lägg inte till eller ta bort elever i Excel-filen."
    assert sheet["C2"].protection.locked is False
    assert sheet["G2"].protection.locked is False
    assert sheet["B2"].protection.locked is True
    assert sheet.protection.sheet is True


@pytest.mark.unit
def test_renderer_links_presentation_sheet_to_hidden_helper_rows():
    workbook = _render_workbook()
    share_sheet = workbook["Dela och exportera"]
    helper_sheet = workbook["_PresentationData"]

    assert share_sheet["A1"].value == "Gruppindelning"
    assert share_sheet["A2"].value == "Klass 7A"
    assert share_sheet["A3"].value == "Skapad 2026-03-26 12:34"
    assert share_sheet["Z5"].value == "='_PresentationData'!Z2"
    assert share_sheet["AA5"].value == "='_PresentationData'!AA2"
    assert share_sheet["B5"].value.startswith('=IF($Z5="heading"')
    assert helper_sheet["B5"].value == (
        '=IF(A5="","",IFERROR(INDEX(\'Redigera grupper\'!$G$2:$G$3,'
        "MATCH(A5,'Redigera grupper'!$F$2:$F$3,0)),\"\"))"
    )
    assert helper_sheet["E5"].value == '=IF(OR(A5="",B5=""),0,1)'
    assert helper_sheet["P3"].value == (
        '=IF(N3="","",1+COUNTIFS($O$2:$O$3,"<"&O3)+COUNTIFS($O$2:$O$3,O3,$N$2:$N$3,"<"&N3))'
    )
    assert helper_sheet["T3"].value == '=IF(R3="","",COUNTIF($H$2:$H$5,R3))'
    assert helper_sheet["U3"].value == '=IF(R3="","",IF(ROW()=2,1,U2+T2+3))'
    assert helper_sheet["M3"].value == (
        '=IF(H3="","",IFERROR(INDEX($U$2:$U$3,MATCH(H3&"|"&I3,$W$2:$W$3,0))+L3+1,""))'
    )
    assert str(share_sheet.page_setup.paperSize) == share_sheet.PAPERSIZE_A4
    assert share_sheet.page_setup.orientation == share_sheet.ORIENTATION_PORTRAIT
    assert share_sheet.page_setup.fitToWidth == 1
    assert share_sheet.page_setup.fitToHeight == 0
    assert share_sheet.print_title_rows == "$1:$3"
