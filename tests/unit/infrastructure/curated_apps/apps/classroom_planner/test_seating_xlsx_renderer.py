"""Unit tests for the classroom-planner seating XLSX renderer.

Purpose:
    Guard the teacher-facing workbook shape so the seating export preserves the
    classroom's spatial layout instead of collapsing it into a coordinate list.

Relationships:
    - Exercises `SeatingXlsxRenderer`.
    - Builds the workbook via the seating XLSX view-model projection.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Sequence
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from skriptoteket.application.curated_apps.classroom_planner.exports import (
    seating_xlsx_view_model,
)
from skriptoteket.domain.curated_apps.classroom_planner.models import (
    ClassroomPlannerWorkspace,
    DraftHistoryStatus,
    PlanDraft,
    PlanDraftKind,
    PlanDraftStatus,
    RoomTemplate,
    Roster,
    Seat,
    SeatAssignment,
    Student,
)
from skriptoteket.infrastructure.curated_apps.apps.classroom_planner.seating_xlsx_renderer import (
    SeatingXlsxRenderer,
)


def _workspace(
    *,
    seats: Sequence[Seat] | None = None,
    seat_assignments: Sequence[SeatAssignment] | None = None,
) -> ClassroomPlannerWorkspace:
    now = datetime(2026, 3, 24, tzinfo=timezone.utc)
    owner_user_id = uuid4()
    roster_id = uuid4()
    template_id = uuid4()
    draft_id = uuid4()

    return ClassroomPlannerWorkspace(
        draft=PlanDraft(
            id=draft_id,
            owner_user_id=owner_user_id,
            roster_id=roster_id,
            draft_kind=PlanDraftKind.SEATING,
            template_id=template_id,
            status=PlanDraftStatus.ACTIVE,
            revision=1,
            last_opened_at=now,
            created_at=now,
            updated_at=now,
        ),
        roster=Roster(
            id=roster_id,
            owner_user_id=owner_user_id,
            name="Klass 7A",
            students=[
                Student(id="student-1", display_name="Ada Lovelace"),
                Student(id="student-2", display_name="Linus Torvalds"),
                Student(id="student-3", display_name="Grace Hopper"),
                Student(id="student-4", display_name="Margaret Hamilton"),
            ],
            created_at=now,
            updated_at=now,
        ),
        template=RoomTemplate(
            id=template_id,
            owner_user_id=owner_user_id,
            name="Sal A",
            seats=list(
                seats
                if seats is not None
                else (
                    Seat(id="seat-1", x=0, y=0),
                    Seat(id="seat-2", x=96, y=0),
                    Seat(id="seat-3", x=288, y=0),
                    Seat(id="seat-4", x=0, y=192),
                )
            ),
            fixtures=[],
            created_at=now,
            updated_at=now,
        ),
        seat_assignments=list(
            seat_assignments
            if seat_assignments is not None
            else (
                SeatAssignment(student_id="student-1", seat_id="seat-1"),
                SeatAssignment(student_id="student-2", seat_id="seat-3"),
                SeatAssignment(student_id="student-3", seat_id="seat-4"),
            )
        ),
        history_status=DraftHistoryStatus(can_undo=False, can_redo=False),
    )


def _render_workbook(
    *,
    workspace: ClassroomPlannerWorkspace | None = None,
):
    view_model = seating_xlsx_view_model.build_seating_xlsx_view_model(
        workspace=workspace or _workspace()
    )
    workbook_bytes = SeatingXlsxRenderer().render(view_model=view_model)
    return load_workbook(BytesIO(workbook_bytes))


@pytest.mark.unit
def test_renderer_uses_single_sheet_and_open_order():
    workbook = _render_workbook()

    assert workbook.sheetnames == ["Sittplacering"]
    assert workbook.active.title == "Sittplacering"


@pytest.mark.unit
def test_renderer_preserves_spatial_layout_with_empty_seats_and_gaps():
    workbook = _render_workbook()
    sheet = workbook["Sittplacering"]

    assert sheet["A1"].value == "Ada Lovelace"
    assert sheet["B1"].value is None
    assert sheet["D1"].value == "Linus Torvalds"
    assert sheet["A3"].value == "Grace Hopper"
    assert sheet["C1"].value is None
    assert sheet["A2"].value is None
    assert sheet["B1"].border.left.style == "thin"
    assert sheet["B1"].fill.fgColor.rgb == "00F3F4F6"
    assert sheet["A1"].comment is not None
    assert sheet["A1"].comment.text == "plats-1"
    assert sheet["B1"].comment is not None
    assert sheet["B1"].comment.text == "plats-2"
    assert sheet.column_dimensions["C"].width == sheet.column_dimensions["B"].width
    assert sheet.row_dimensions[1].height == sheet.row_dimensions[2].height


@pytest.mark.unit
def test_renderer_lists_unplaced_students_below_grid_and_sets_print_layout():
    workbook = _render_workbook()
    sheet = workbook["Sittplacering"]

    assert sheet["A6"].value == "Ej placerade elever"
    assert sheet["A7"].value == "Elevnamn"
    assert sheet["B7"].fill.patternType is None
    assert sheet["A8"].value == "Margaret Hamilton"
    assert str(sheet.page_setup.paperSize) == sheet.PAPERSIZE_A4
    assert sheet.page_setup.orientation == sheet.ORIENTATION_LANDSCAPE
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 0


@pytest.mark.unit
def test_renderer_skips_unplaced_section_when_every_student_has_a_seat():
    workbook = _render_workbook(
        workspace=_workspace(
            seat_assignments=(
                SeatAssignment(student_id="student-1", seat_id="seat-1"),
                SeatAssignment(student_id="student-2", seat_id="seat-2"),
                SeatAssignment(student_id="student-3", seat_id="seat-3"),
                SeatAssignment(student_id="student-4", seat_id="seat-4"),
            )
        )
    )
    sheet = workbook["Sittplacering"]

    assert sheet["A5"].value is None
    assert sheet["A6"].value is None


@pytest.mark.unit
def test_renderer_starts_unplaced_section_at_a1_when_template_has_no_seats():
    workbook = _render_workbook(
        workspace=_workspace(
            seats=(),
            seat_assignments=(),
        )
    )
    sheet = workbook["Sittplacering"]

    assert sheet["A1"].value == "Ej placerade elever"
    assert sheet["A2"].value == "Elevnamn"
    assert sheet["A3"].value == "Ada Lovelace"
